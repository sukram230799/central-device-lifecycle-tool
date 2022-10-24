from datetime import datetime

from Generics.GenericExcelHandler import GenericExcelHandler
from openpyxl.styles.numbers import FORMAT_DATE_DATETIME


class FirmwareExcelHandler(GenericExcelHandler):

    def __init__(self, filename, uuid=None) -> None:
        self.name = 'Firmware'
        self.empty_state = {}
        super().__init__(filename)

    def init_worksheet(self):
        self.worksheet.cell(row=1, column=1, value='Serial')
        self.worksheet.cell(row=1, column=2, value='Status')
        self.worksheet.cell(row=1, column=3, value='Date')

        self.last_row = 2

    def update_status(self, serial, status):
        excel_row, device_info = self.update_state_internal(
            serial=serial, state={'status': status})

        self.worksheet.cell(row=excel_row, column=1, value=str(serial))
        self.worksheet.cell(row=excel_row,
                            column=2,
                            value=str(device_info['status']))
        self.worksheet.cell(row=excel_row, column=3, value=datetime.now(
        )).number_format = FORMAT_DATE_DATETIME  # type: ignore

        self.save()
