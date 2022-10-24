from datetime import datetime
from typing import TypedDict

from Generics.GenericExcelHandler import GenericExcelHandler
from openpyxl.styles.numbers import FORMAT_DATE_DATETIME


class DecomissionState(TypedDict):
    status: str
    excel_row: str
    deleted_on: str
    unsubscribed_on: str


class DecomissionExcelHandler(GenericExcelHandler):

    def __init__(self, filename, uuid=None) -> None:
        self.name = 'Decomission'
        self.empty_state = {
            'status': '',
            'deleted_on': '',
            'unsubscribed_on': ''
        }
        super().__init__(filename)

    def init_worksheet(self):
        self.worksheet.cell(row=1, column=1, value='Serial')
        self.worksheet.cell(row=1, column=2, value='Status')
        self.worksheet.cell(row=1, column=3, value='Last State update')
        self.worksheet.cell(row=1, column=4, value='Deleted on')
        self.worksheet.cell(row=1, column=5, value='Unsubscribed on')

        self.last_row = 2

    def update_status(self, serial, state: dict):
        excel_row, state = self.update_state_internal(serial=serial,
                                                            state=state)

        self.worksheet.cell(row=excel_row, column=1, value=str(serial))
        self.worksheet.cell(row=excel_row,
                            column=2,
                            value=state['status'])
        self.worksheet.cell(row=excel_row, column=3, value=datetime.now(
        )).number_format = FORMAT_DATE_DATETIME  # type: ignore
        self.worksheet.cell(row=excel_row,
                            column=4,
                            value=state['deleted_on'])
        self.worksheet.cell(row=excel_row,
                            column=5,
                            value=state['unsubscribed_on'])

        self.save()
