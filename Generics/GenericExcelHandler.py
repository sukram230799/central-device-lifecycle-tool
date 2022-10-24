import os
import os.path
import typing
from abc import ABC, abstractmethod
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class GenericExcelHandler(ABC):

    def __init__(self, filename, uuid=None) -> None:
        if not self.name:
            self.name = ''
        if os.path.exists(filename):
            if os.path.isdir(filename):
                filename = os.path.join(filename, self.filename_gen())
            if os.path.isfile(filename):
                self.workbook = load_workbook(filename=filename)
                self.worksheet = typing.cast(
                    Worksheet,
                    self.workbook.create_sheet(self.sheetname_gen(),
                                               index=0))  # type: ignore
                self.workbook.save(filename)
            else:
                self.workbook = Workbook(write_only=False)
                self.worksheet = self.workbook.active
                self.worksheet.title = self.sheetname_gen()
        else:
            self.workbook = Workbook(write_only=False)
            self.worksheet = self.workbook.active
            self.worksheet.title = self.sheetname_gen()

        # self.worksheet = self.workbook.active
        # self.worksheet.title = ExcelHandler.sheetname_gen()

        self.filename = filename
        self.workbook.save(self.filename)

        if not self.empty_state:
            self.empty_state = {}

        self.last_row = 2
        self.init_worksheet()

        self.device_states = {
        }  # f.e. {'SN01': {'state': 'Updated', 'excel_row': 1}}

    def filename_gen(self):
        return datetime.now().strftime(f"%Y-%m-%d_%Hh%Mm%Ss_{self.name}.xlsx")

    def sheetname_gen(self):
        return datetime.now().strftime("%Y-%m-%d_%Hh%Mm%Ss")

    def update_state_internal(self, serial,
                              state: dict) -> typing.Tuple[int, dict]:
        if serial not in self.device_states:
            self.device_states[serial] = self.empty_state

        device_state = self.device_states[serial]
        device_state = {**device_state, **state}

        excel_row = ''
        if 'excel_row' in device_state.keys():
            excel_row = device_state['excel_row']
        else:
            excel_row = self.last_row
            device_state['excel_row'] = excel_row
            self.last_row += 1

        self.device_states[serial] = device_state

        return excel_row, device_state

    @abstractmethod
    def init_worksheet(self):
        pass

    def close(self):
        self.workbook.save(filename=self.filename)
        return self.workbook.close()

    def get_filename(self):
        return os.path.basename(self.filename)

    def delete(self):
        self.close()
        return os.remove(self.filename)

    def save(self):
        return self.workbook.save(self.filename)
